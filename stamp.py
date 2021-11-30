# -*- coding; utf-8 -*-

import random
import copy
import openpyxl

num_dimension=11    #駅の数
max_time = 450     #時間の上限[分]

ikebukuro =[] #池袋方向の時刻表
hanno = [] #飯能方向の時刻表
toshimaen = []  #豊島線
honkawa = []    #本川越方向
seibushinjuku = []  #西武新宿方向

#乗車時間
ih_ir =[]   #入間
ih_tz =[]   #所沢
ih_sk =[]   #石神井公園
ih_nm=[]    #練馬
ih_ib=[]    #池袋
hs_st=[]    #新所沢
hs_tz=[]    #所沢
hs_tn=[]    #田無
hs_um=[]    #宮
hs_tb=[]    #高田馬場
hs_ss=[]    #西武新宿

pop = []    #全個体のスタンプ取得順番情報
offspring = []  #子個体のスタンプ取得順番情報
sum_time = []    #全個体の要する時間
size_pop = 20   #個体数
max_generation = 100    #繰り返し回数の上限

#列のリストを作成
def make_timeline(sheet):
    list1=[]
    list2=[]
    for col in list(sheet.columns)[1]:
             list1.append(col.value)

    for col in list(sheet.columns)[2]:
             list2.append(col.value)

    return list1,list2

#行のリストを作成
def make_timerow(sheet):
    l=[[0 for j in range(sheet.max_column)] for i in range(sheet.max_row)]
    for i in range(sheet.max_row):
        j=0
        for col in list(sheet.rows)[i]:
            l[i][j]=col.value
            j+=1
    l.pop(0)

    return l

 #データの読み取り
def read_data_i(file_name):  #池袋方向
    wb = openpyxl.load_workbook(file_name)

    #シートを取得
    s1 = wb['musa']
    s2 = wb['iruma']
    s3 = wb['toko']
    s4 = wb['shaku']
    s5 = wb['neri']

    #リストを作る
    l=[]
    x,y=make_timeline(s1)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s2)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s3)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s4)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s5)
    l.append(x)
    l.append(y)

    return l

def read_data_h(file_name):  #飯能方向
    wb = openpyxl.load_workbook(file_name)

    #シートを取得
    s1 = wb['musa']
    s2 = wb['toko']
    s3 = wb['shaku']
    s4 = wb['neri']
    s5 = wb['ike']

    #リストを作る
    l=[]
    x,y=make_timeline(s1)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s2)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s3)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s4)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s5)
    l.append(x)
    l.append(y)

    return l

def read_data_t(file_name):  #豊島園方向
    wb = openpyxl.load_workbook(file_name)

    #シートを取得
    s1 = wb['to']
    s2 = wb['ne']

    #リストを作る
    l=[]
    x,y=make_timeline(s1)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s2)
    l.append(x)
    l.append(y)

    return l

def read_data_hk(file_name):  #本川越方向
    wb = openpyxl.load_workbook(file_name)

    #シートを取得
    s1 = wb['toko']
    s2 = wb['tana']
    s3 = wb['miya']
    s4 = wb['taka']
    s5 = wb['shin']

    #リストを作る
    l=[]
    x,y=make_timeline(s1)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s2)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s3)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s4)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s5)
    l.append(x)
    l.append(y)

    return l

def read_data_ss(file_name):  #西武新宿方向
    wb = openpyxl.load_workbook(file_name)

    #シートを取得
    s1 = wb['toko']
    s2 = wb['shint']
    s3 = wb['tana']
    s4 = wb['miya']
    s5 = wb['takada']

    #リストを作る
    l=[]
    x,y=make_timeline(s1)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s2)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s3)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s4)
    l.append(x)
    l.append(y)
    x, y = make_timeline(s5)
    l.append(x)
    l.append(y)

    return l
#乗車時間の読み取り
def read_data_ih(file_name):    #西武池袋線
    wb = openpyxl.load_workbook(file_name)

    # シートを取得
    s1 = wb['iruma']
    s2 = wb['tokorozawa']
    s3 = wb['shaku']
    s4 = wb['nerima']
    s5 = wb['ikebukuro']

    return make_timerow(s1), make_timerow(s2), make_timerow(s3), make_timerow(s4), make_timerow(s5)

#乗車時間の読み取り
def read_data_hs(file_name):
    wb = openpyxl.load_workbook(file_name)

    # シートを取得
    s1 = wb['shintoko']
    s2 = wb['toko_h']
    s3 = wb['tanashi']
    s4 = wb['nomiya']
    s5 = wb['takada']
    s6 = wb['s_shin']

    return make_timerow(s1), make_timerow(s2), make_timerow(s3), make_timerow(s4), make_timerow(s5),make_timerow(s6)

def initialize():   #ランダムに個体集団を生成する
    for i in range(size_pop):
        #1~11をランダムに並び替えたリストをpopに追加
        print(random.sample(range(12),k=12))
        l=random.sample(range(12),k=12)
        l.remove(0)
        pop.append(l)

def crossover(pop_current): #交叉して子集団を生成する
    parents =[] #親個体集団
    offspring = []  #子集団（予定)

    for i in range(int(size_pop / 2)):
        while True:
            # 交叉させる親個体をランダム選択(要素番号)
            i1 = random.randint(0,size_pop-1)
            i2 = random.randint(0,size_pop-1)
            if i1 != i2:
                break
        #選択した２つをparentsリストに追加
        parents.append(pop_current[i1])
        parents.append(pop_current[i2])
        #マスクを生成する(0,1)で表現
        mask = [random.randint(0,1) for i in range(num_dimension)]

        #o1,o2:交叉によって生成される子個体
        o1 = pop_current[i1]
        o2 = pop_current[i2]

        #一様順序交叉を適用する
        # 入れ替える所の要素を記録する
        m1=[]
        m2=[]
        for d in range(num_dimension):
            if mask[d] == 1:
                m1.append(parents[i*2][d])
                m2.append(parents[i*2+1][d])
        #入れ替える要素番号を調べる
        n1=[]
        n2=[]
        for d in m2:
            n1.append(parents[i*2].index(d))
        for d in m1:
            n2.append(parents[i * 2+1].index(d))

        #   入れ替えて、子個体生成する
        for d in range(len(m1)):
            o1[n1[d]] = m2[d]
            o2[n2[d]] = m1[d]

        #生成した子個体をoffspringに追加
        offspring.append(o1)
        offspring.append(o2)

    #子個体を返す
    return offspring

def mutation(pop_current):  #突然変異を起こす
    offspring = copy.deepcopy(pop_current)
    for i in range(size_pop):
        #入れ替える2点(要素番号)をランダムに選ぶ
        num = [random.randint(0,10),random.randint(0,10)]
        #選んだ2点の要素を入れ替える
        offspring[i][num[0]], offspring[i][num[1]] = offspring[i][num[1]],offspring[i][num[0]]

    #突然変異を適用した子個体を返す
    return offspring

def best_selection(t,cur,nxt,sheet,num):    #電車の来る時間と乗車時間を返す
    time=t
    st=1000    #来る時間
    zyosha=120    #乗車時間
    c_tt=[] #乗車時刻の候補
    o_tt=[] #乗車時間の候補

    if num==1 or num==2:  #池袋方向:1 #飯能方向:2
        #出発時間の候補
        if num==1:  #池袋方向
            for i in range(30): #30分間の電車を調べる
              if sheet[2*cur].count(time+i) == 1:
                 c_tt.append([time+i,int(sheet[2*cur+1][sheet[2*cur].index(time+i)])])

        else:   #num==2
            for i in range(30):
              if sheet[2*cur-2].count(time+i) == 1:
                 c_tt.append([time+i,int(sheet[2*cur-1][sheet[2*cur-2].index(time+i)])])

        #乗車時間を求める
        for d in c_tt:
            i=d[1]
            #今いる場所で使う表が違う
            if cur==1:
                o_tt.append(ih_ir[i][nxt])
            elif cur==2:
                o_tt.append(ih_tz[i][nxt])
            elif cur == 3:
                o_tt.append(ih_sk[i][nxt])
            elif cur == 4:
                o_tt.append(ih_nm[i][nxt])
            elif cur==5:
                o_tt.append(ih_ib[i][nxt])

        #一番良いもの決定する
        l_t=1000    #到着時間
        i = 0
        for d in o_tt:
            #到着時間=出発時刻+乗車時間
            if not d ==0 and l_t > d + c_tt[i][0]:
                l_t=d+c_tt[i][0]
                zyosha=d
                st=c_tt[i][0]
            i+=1
        return st, zyosha

    elif num==3:    #豊島園
        #各駅停車しかない(3分)
        if(nxt==6):#練馬から豊島園
            for l in range(60):
                if sheet[2].count(time)==1:
                    break
                else:
                    time +=1
            #出発時間
            st=time
            #乗車時間
            zyosha=3

        else:   #豊島園から練馬
            for l in range(60):
                if sheet[0].count(time)==1:
                    break
                else:
                    time += 1
            #出発時間
            st = time
            #乗車時間
            zyosha = 3

        return st,zyosha

    elif num==4 :    #本川越方向:4
        # 出発時間の候補
        if cur == 2:    #今所沢
            for i in range(30):
                if sheet[0].count(time+i) == 1:
                    c_tt.append([time+i, sheet[1][sheet[0].index(time+i)]])

        else:# num==4:  # cur=8,9,10,11
            for i in range(30):
                if sheet[2 * cur - 14].count(time+i) == 1:
                    c_tt.append([time+i, sheet[2 * cur - 13][sheet[2 * cur - 14].index(time+i)]])

        # 乗車時間を求める
        for d in c_tt:
            #ラベル30~34
            am = d[1] % 10
            if cur == 2:
                o_tt.append(hs_tz[am][nxt-5])
            elif cur == 8:
                if nxt == 2:
                    o_tt.append(hs_tn[am][1])
                else:
                    o_tt.append(hs_tn[am][nxt-5])
            elif cur == 9:
                if nxt == 2:
                    o_tt.append(hs_tn[am][1])
                else:
                    o_tt.append(hs_um[am][nxt-5])
            elif cur == 10:
                if nxt == 2:
                    o_tt.append(hs_tb[am][1])
                else:
                    o_tt.append(hs_tb[am][nxt-5])
            else:   #cur==11
                if nxt == 2:
                    o_tt.append(hs_ss[am][1])
                else:
                    o_tt.append(hs_ss[am][nxt-5])

        # 一番良いもの決定する
        l_t = 5000  # 到着時間
        i=0
        for d in o_tt:
            if not d == 0 and l_t > d + c_tt[i][0]:
                l_t=d+c_tt[i][0]
                zyosha = d
                st = c_tt[i][0]
            i+=1
        return st, zyosha

    else:   #num==5 #西武新宿:5 本川越の構造と同じ
        #出発時間の候補
        if cur==2:  #所沢
            for i in range(30):
                if sheet[0].count(time+i) == 1:
                    c_tt.append([time+i, sheet[1][sheet[0].index(time+i)]])
        else:   #cur>=7
            for i in range(30):
                if sheet[2*cur-12].count(time+i) == 1:
                    c_tt.append([time+i, sheet[2*cur-11][sheet[2*cur-12].index(time+i)]])

        # 乗車時間を求める
        for d in c_tt:
            i = d[1] % 10
            if cur == 7:
                if nxt==2:
                    o_tt.append(hs_st[i][1])
                else:
                    o_tt.append(hs_st[i][nxt-5])
            elif cur == 2:
                o_tt.append(hs_tz[i][nxt-5])
            elif cur == 8:
                if nxt==2:
                    o_tt.append(hs_tn[i][1])
                else:
                    o_tt.append(hs_tn[i][nxt-5])
            elif cur == 9:
                if nxt==2:
                     o_tt.append(hs_um[i][1])
                else:
                    o_tt.append(hs_um[i][nxt-5])
            elif cur == 10:
                if nxt==2:
                    o_tt.append(hs_tb[i][1])
                else:
                    o_tt.append(hs_tb[i][nxt-5])

        # 一番良いもの決定する
        l_t = 5000  # 到着時間
        i = 0
        for d in o_tt:
            if not d == 0 and l_t > d + c_tt[i][0]:
                l_t = d + c_tt[i][0]
                zyosha = d
                st = c_tt[i][0]
                i += 1

        return st, zyosha

def evaluate(pop_current,N):  #目的関数の計算
    fit_current = []    #計算結果を保存
    for i in range(size_pop):   #i番目の個体を選択
        order=pop_current[i]    #順番配列
        #武蔵藤沢を出発時間
        total_t =30 #時刻表が9時から始まり、出発時間は9:30のため
        nxt=order.index(1) + 1  #次
        cur=0   #現在地
        #次の目的地によって場合分け。今は必要ない
        #武蔵藤沢から始めの駅:nxt=1
        if nxt==1:  #入間
            while not hanno[0].count(total_t) == 1:
                total_t +=1
                cur=nxt
            total_t+=7

        elif nxt==2:    #所沢
            while not ikebukuro[0].count(total_t) == 1:
                total_t += 1
            total_t += 12
            cur=nxt

        elif nxt>2 and nxt <=5: #池袋寄りの駅
            #所沢に行く
            while not ikebukuro[0].count(total_t) == 1:
                total_t += 1
            total_t += 12
            cur=2
            #nxtに移動
            t_d, t_t = best_selection(total_t, cur, nxt, ikebukuro, 1)
            total_t=t_d+t_t
            cur=nxt

        elif nxt == 6:  #豊島園
            # 所沢に行く
            while not ikebukuro[0].count(total_t) == 1:
                total_t += 1
            total_t += 12
            cur=2
            # 練馬に移動
            t_d, t_t = best_selection(total_t, cur, 4, ikebukuro, 1)
            total_t = t_d + t_t+5
            cur=4
            # 豊島園行きに乗車
            t_d, t_t = best_selection(total_t, cur, nxt, toshimaen, 3)
            total_t = t_d + t_t
            cur=nxt
        else:
            # 所沢に行く
            while not ikebukuro[0].count(total_t) == 1:
                total_t += 1
            total_t += 10+5
            cur = 2
            if nxt == 7:    #新所沢
                t_d, t_t = best_selection(total_t, cur, nxt, honkawa, 4)
                total_t = t_d + t_t
                cur=nxt
            else:#西武新宿方向
                t_d, t_t = best_selection(total_t, cur, nxt, seibushinjuku, 5)
                total_t = t_d + t_t
                cur=nxt

        #スタンプ取得
        total_t += 10

        #2~Nまでの所要時間を求める
        for intint in range(N-1):
            nxt = order.index(intint+2) + 1

            if nxt<=5: #西武池袋線
                if cur<nxt:
                    t_d,t_t = best_selection(total_t,cur,nxt,ikebukuro,1)
                    total_t=t_d+t_t
                    cur=nxt
                elif cur>nxt and cur <=5:
                    t_d, t_t = best_selection(total_t, cur, nxt, hanno,2)
                    total_t = t_d + t_t
                    cur=nxt
                elif cur == 6:
                    #練馬乗り換え
                    t_d, t_t = best_selection(total_t, cur, 4, toshimaen,3)
                    total_t = t_d + t_t+5
                    cur=4
                    #nxtに行く
                    if nxt==4:
                        exit
                    elif nxt==5:  #上り
                        t_d, t_t = best_selection(total_t, cur, nxt, ikebukuro,1)
                        total_t = t_d + t_t
                        cur=nxt
                    else:   #下り
                        t_d, t_t = best_selection(total_t, cur, nxt, hanno,2)
                        total_t = t_d + t_t
                        cur=nxt
                else:# cur>=7:
                    #所沢乗り換え
                    if cur == 7:    #新所沢
                        t_d, t_t = best_selection(total_t, cur, 2, seibushinjuku,5)
                        total_t = t_d + t_t+5
                        cur=2
                    else:
                        t_d, t_t = best_selection(total_t, cur, 2, honkawa,4)
                        total_t = t_d + t_t+5
                        cur=2
                     #nxtに行く
                    if nxt == 2:
                        exit
                    elif nxt>2:  #上り
                        t_d, t_t = best_selection(total_t, cur, nxt, ikebukuro,1)
                        total_t = t_d + t_t
                        cur=nxt
                    else:   #下り
                        t_d, t_t = best_selection(total_t, cur, nxt, hanno,2)
                        total_t = t_d + t_t
                        cur=nxt

            elif nxt>=7:    #西武新宿線
                if cur>=7 or cur ==2:   #同じ路線上
                    if cur>nxt or (cur==2 and nxt==7) or (nxt==2 and cur>7):   #本川越方向
                        t_d, t_t = best_selection(total_t, cur, nxt, honkawa,4)
                        total_t = t_d + t_t
                        cur=nxt
                    else:   #西武新宿方向
                        t_d, t_t = best_selection(total_t, cur, nxt, seibushinjuku,5)
                        total_t = t_d + t_t
                        cur=nxt
                else:# cur <=6:   #所沢で乗り換えあり
                    if cur == 6:   #豊島園にいる時は練馬に移動
                        t_d, t_t = best_selection(total_t, cur, 4, toshimaen,3)
                        total_t = t_d + t_t+5
                        cur = 4

                    #所沢に行く
                    if cur==1:
                        t_d, t_t = best_selection(total_t, cur, 2, ikebukuro,1)
                        total_t = t_d + t_t+5
                        cur=2
                    else:
                        t_d, t_t = best_selection(total_t, cur, 2, hanno,2)
                        total_t = t_d + t_t+5
                        cur=2
                    #所沢から目的地に行く
                    if nxt == 7:    #新所沢
                        t_d, t_t = best_selection(total_t, cur, nxt, honkawa,4)
                        total_t = t_d + t_t
                        cur=nxt
                    else:   #西武新宿方向
                        t_d, t_t = best_selection(total_t, cur, nxt,seibushinjuku,5 )
                        total_t = t_d + t_t
                        cur=nxt
            else:  #豊島園 nxt==6
                #西武新宿線は所沢で乗り換えする
                if cur ==7:
                    t_d, t_t = best_selection(total_t, cur, 2, seibushinjuku,5)
                    total_t = t_d + t_t+5
                    cur=2
                elif cur >=8:
                    t_d, t_t = best_selection(total_t, cur, 2, honkawa,4)
                    total_t = t_d + t_t+5
                    cur=2

                #練馬で乗り換える
                if cur <= 3:
                    t_d, t_t = best_selection(total_t, cur, 4, ikebukuro,1)
                    total_t = t_d + t_t+5
                    cur=4
                elif cur==5:
                    t_d, t_t = best_selection(total_t, cur, 4, hanno,2)
                    total_t = t_d + t_t+5
                    cur=4

                #豊島園行きに乗車
                t_d, t_t = best_selection(total_t, cur, nxt, toshimaen,3)
                total_t = t_d + t_t
                cur=nxt

            #スタンプを取得する(5分)
            total_t +=5

        #景品を交換する
        if cur==4 or cur==6 or cur==9 or cur== 10:  #交換できない駅にいる
            #所沢駅で交換
            if cur ==6: #豊島園にいるとき練馬に行く
                t_d, t_t = best_selection(total_t, cur, 4, toshimaen,3)
                total_t = t_d + t_t+5
                cur=4

            if cur==4:  #練馬にいる
                t_d, t_t = best_selection(total_t, cur, 2, hanno,1)
                total_t = t_d + t_t
                cur=2
            else:   #西武新宿線にいる
                t_d, t_t = best_selection(total_t, cur, 2, honkawa,4)
                total_t = t_d + t_t
                cur=2
        #交換10分
        total_t += 15

        #最寄り駅に帰る
        if cur==1:  #入間->武蔵藤沢：5分
            #F快速急行は止まらない
            for l in range(30):
                if ikebukuro[2].count(total_t) == 1 and not ikebukuro[3][ikebukuro[2].index(total_t)]== 9:
                    break
                else:
                    total_t +=1

            total_t += 5
        elif cur<=5 and not cur==1: #池袋線上にいる
            #所沢に行く
            if cur<2:
                t_d, t_t = best_selection(total_t, cur, 2, hanno,2)
                total_t = t_d + t_t

            #所沢->武蔵藤沢:10分
            for l in range(30):
                if ikebukuro[2].count(total_t) == 1:
                    v=ikebukuro[3][ikebukuro[2].index(total_t)]
                    if not v==10 and not v==11 and not v==12:   #小手指止まりではない
                        break
                else:
                    total_t +=1
            total_t +=10

        elif cur==7 or cur==8 or cur==11:   #cur=7,8,11
            #所沢で乗り換え
            if cur==7:
                t_d, t_t = best_selection(total_t, cur, 2, seibushinjuku,5)
                total_t = t_d + t_t+5
                cur=2
            else:
                t_d, t_t = best_selection(total_t, cur, 2, honkawa,4)
                total_t = t_d + t_t+5
                cur=2
            # 所沢->武蔵藤沢:10分
            for l in range(30):
                if ikebukuro[2].count(total_t) == 1 :
                    v = ikebukuro[3][ikebukuro[2].index(total_t)]
                    if not v == 10 and not v == 11 and not v == 12:  # 小手指止まりではない
                        break
                else:
                    total_t +=1
        else:
            print("帰れない")

        #合計時間を保存
        fit_current.append(total_t - 30)  #開始時刻がt=30のため

    return fit_current

def select(parent, offspring,fitness_parent,fitness_offspring): #次世代に残す個体の選別
    pop_next =[]
    fit_next = []

    for i in range(int(size_pop /2)):
        cand = []   #個体
        cand.extend([parent[i*2]])
        cand.extend([parent[i *2+1]])
        cand.extend([offspring[i * 2]])
        cand.extend([offspring[i * 2+1]])
        fit = []    #目的関数の値
        fit.append(fitness_parent[i*2])
        fit.append(fitness_parent[i * 2+1])
        fit.append(fitness_offspring[i * 2])
        fit.append(fitness_offspring[i * 2+1])

        indi_min = fit.index(min(fit))  #目的関数が最小なもの
        while True: #ランダムに一つ選ぶ
            indi_rand = random.randint(0,3)
            if indi_rand != indi_min:
                  if fit[indi_rand]<max_time:  #制約条件を満たすもの
                        break

        #目的関数が最小なものとランダムに選んだものを残す
        pop_next.append(cand[indi_min])
        pop_next.append(cand[indi_rand])

        fit_next.append(fit[indi_min])
        fit_next.append(fit[indi_rand])

    return pop_next, fit_next

#リスト全部を表示する
def print_pop(pop_current,fit):
    for i in range(size_pop):
        print("Indi {} : {} {}".format(i, pop_current[i],fit[i]))

#1~Nまでの駅を表示する
def print_eki(pop,fit,num):
    lis=[]
    for i in range(size_pop):
        lis.clear()
        for n in range(num):
            lis.append(pop[i].index(n+1)+1)
        print("Indi {} : {} (Fit:{})".format(i, lis,fit[i]))

#main
#データを読み取る
ikebukuro=read_data_i("i_train_schedule.xlsx")
hanno=read_data_h("h_train_schedule.xlsx")
toshimaen=read_data_t("t_train_schedule.xlsx")
honkawa=read_data_hk("hk_train_schedule.xlsx")
seibushinjuku = read_data_ss("ss_train_schedule.xlsx")
#乗車時間データ
ih_ir,ih_tz,ih_sk,ih_nm,ih_ib,=read_data_ih("i_h.xlsx")
hs_st,hs_tz,hs_tn,hs_um,hs_tb,hs_ss=read_data_hs("h_s.xlsx")

#スタンプ取得数を得る
num_stamp=5

#初期個体集団を生成する
initialize()
print("Initial population:")

#初期個体に対して目的関数を計算する
fitness = evaluate(pop,num_stamp)

#初期個体の表示
print_pop(pop,fitness)

#繰り返す
for g in range(max_generation):
    #親集団popに交叉を適用し、子個体集団を作成
    offspring = crossover(pop)
    #生成した子個体集団に突然変異を適用
    offspring = mutation(offspring)
    #生成した子個体集団を評価
    fitness_offspring = evaluate(offspring,num_stamp)
    #親子両集団から次世代に残す個体を選択
    pop,fitness = select(pop,offspring,fitness,fitness_offspring)
    #目的関数が最小な個体を調べる
    indi_best = fitness.index(min(fitness))
    print_eki(pop,fitness,num_stamp)
    print("Best indi at gen {}:{} (Fitness: {}){}"\
        .format(g,indi_best,fitness[indi_best],pop[indi_best]))